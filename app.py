import streamlit as st
from llama_index.core import StorageContext, load_index_from_storage, VectorStoreIndex, SimpleDirectoryReader, ChatPromptTemplate
from llama_index.llms.huggingface import HuggingFaceInferenceAPI
from dotenv import load_dotenv
from llama_index.embeddings.huggingface import HuggingFaceEmbedding
from llama_index.core import Settings
import os
import base64
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# Load environment variables
load_dotenv()

# Configure the Llama index settings
Settings.llm = HuggingFaceInferenceAPI(
    model_name="google/gemma-1.1-7b-it",
    tokenizer_name="google/gemma-1.1-7b-it",
    context_window=3000,
    token=os.getenv("HF_TOKEN"),
    max_new_tokens=512,
    generate_kwargs={"temperature": 0.1},
)
Settings.embed_model = HuggingFaceEmbedding(
    model_name="BAAI/bge-small-en-v1.5"
)

# Define the directory for persistent storage and data
PERSIST_DIR = "./db"
DATA_DIR = "data"

# Ensure data directory exists
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(PERSIST_DIR, exist_ok=True)

def displayPDF(file):
    with open(file, "rb") as f:
        base64_pdf = base64.b64encode(f.read()).decode('utf-8')
    pdf_display = f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="100%" height="600" type="application/pdf"></iframe>'
    st.markdown(pdf_display, unsafe_allow_html=True)

def scrape_url(url):
    try:
        response = requests.get(url)
        response.raise_for_status()  # Ensure we notice bad responses
        soup = BeautifulSoup(response.content, 'html.parser')
        text = soup.get_text()
        # Save the text content to a file for processing
        text_file_path = os.path.join(DATA_DIR, "scraped_content.txt")
        with open(text_file_path, "w") as file:
            file.write(text)
        return text_file_path
    except requests.RequestException as e:
        st.error(f"Error fetching the URL: {e}")
        return None

def process_text_file(file):
    text_file_path = os.path.join(DATA_DIR, "uploaded_text.txt")
    with open(file, "r") as f:
        text = f.read()
    with open(text_file_path, "w") as f:
        f.write(text)
    return text_file_path

def process_excel_file(file):
    text_file_path = os.path.join(DATA_DIR, "uploaded_excel.txt")
    wb = load_workbook(file)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append("\t".join([str(cell) for cell in row if cell is not None]))
    text = "\n".join(data)
    with open(text_file_path, "w") as f:
        f.write(text)
    return text_file_path

def data_ingestion():
    documents = SimpleDirectoryReader(DATA_DIR).load_data()
    storage_context = StorageContext.from_defaults()
    index = VectorStoreIndex.from_documents(documents)
    index.storage_context.persist(persist_dir=PERSIST_DIR)

def handle_query(query):
    storage_context = StorageContext.from_defaults(persist_dir=PERSIST_DIR)
    index = load_index_from_storage(storage_context)
    chat_text_qa_msgs = [
        (
            "user",
            """You are a Q&A assistant named LORAIN, created by LOR Technologies. You have a specific response programmed for when users specifically ask about your creator, Suriya. The response is: "I was created by Suriya, an enthusiast in Artificial Intelligence. He is dedicated to solving complex problems and delivering innovative solutions. With a strong focus on machine learning, deep learning, Python, generative AI, NLP, and computer vision, Suriya is passionate about pushing the boundaries of AI to explore new possibilities." For all other inquiries, your main goal is to provide answers as accurately as possible, based on the instructions and context you have been given. If a question does not match the provided context or is outside the scope of the document, kindly advise the user to ask questions within the context of the document.
            Context:
            {context_str}
            Question:
            {query_str}
            """
        )
    ]
    text_qa_template = ChatPromptTemplate.from_messages(chat_text_qa_msgs)
    
    query_engine = index.as_query_engine(text_qa_template=text_qa_template)
    answer = query_engine.query(query)
    
    if hasattr(answer, 'response'):
        return answer.response
    elif isinstance(answer, dict) and 'response' in answer:
        return answer['response']
    else:
        return "Sorry, I couldn't find an answer. You can either elect to contiinue this chat or be diverted to a Live Agent"

# Streamlit app initialization
st.title("LOR Technologies Retrieval-Augmented Generation")
st.markdown("start chat ...🚀")

if 'messages' not in st.session_state:
    st.session_state.messages = [{'role': 'assistant', "content": 'Hello! Upload a PDF, Excel, text file, or provide a URL and ask me anything about its content.'}]

with st.sidebar:
    st.title("Menu:")
    # Upload PDF, Excel, or Text File
    uploaded_file = st.file_uploader("Upload your PDF, Excel, or Text Files", type=["pdf", "xlsx", "txt"])
    url_input = st.text_input("Or enter a URL to scrape content from")
    
    if st.button("Submit & Process"):
        with st.spinner("Processing..."):
            if uploaded_file is not None:
                file_extension = uploaded_file.name.split(".")[-1].lower()
                if file_extension == "pdf":
                    filepath = os.path.join(DATA_DIR, "saved_pdf.pdf")
                    with open(filepath, "wb") as f:
                        f.write(uploaded_file.getbuffer())
                    displayPDF(filepath)  # Display the uploaded PDF
                elif file_extension == "txt":
                    filepath = process_text_file(uploaded_file)
                elif file_extension == "xlsx":
                    filepath = process_excel_file(uploaded_file)
                data_ingestion()  # Process the uploaded file
                st.success("File processed successfully!")
            elif url_input:
                file_path = scrape_url(url_input)
                if file_path:
                    data_ingestion()  # Process text content scraped from URL
                    st.success("URL content processed successfully!")
                else:
                    st.error("Failed to process URL content.")
            else:
                st.warning("Please upload a file or enter a URL.")

user_prompt = st.chat_input("Ask me anything about the content:")
if user_prompt:
    st.session_state.messages.append({'role': 'user', "content": user_prompt})
    response = handle_query(user_prompt)
    st.session_state.messages.append({'role': 'assistant', "content": response})

for message in st.session_state.messages:
    with st.chat_message(message['role']):
        st.write(message['content'])